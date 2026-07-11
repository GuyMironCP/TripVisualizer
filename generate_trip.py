# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def border(color="FFB0BEC5"):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def cell_style(ws, addr, value, bg, font_color="FF212121", bold=False, size=9, wrap=True, center=True):
    c = ws[addr]
    c.value = value
    c.fill = fill(bg)
    c.font = Font(name='Arial', bold=bold, color=font_color, size=size)
    halign = 'center' if center else 'left'
    c.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=wrap)
    c.border = border()
    return c

# ── Color palette ──────────────────────────────────────────────────────────────
C_GRAY   = "FFB0BEC5"   # flights / airport / packing
C_BLUE   = "FF90CAF9"   # car rental / return
C_ORANGE = "FFFFCC80"   # driving
C_GREEN  = "FFA5D6A7"   # nature / hiking / lakes
C_YELLOW = "FFFFF59D"   # attractions / parks / adventure

C_HDR_DARK  = "FF1A237E"  # title row
C_HDR_H1    = "FF1565C0"  # hotel 1 header
C_HDR_H2    = "FF6A1B9A"  # hotel 2 header
C_HDR_END   = "FFC62828"  # return day header
C_DAY_H1    = "FF1976D2"
C_DAY_H2    = "FF7B1FA2"
C_DAY_END   = "FFD32F2F"
C_TIME_COL  = "FF37474F"
C_EMPTY_H1  = "FFE3F2FD"
C_EMPTY_H2  = "FFF3E5F5"
C_EMPTY_END = "FFFCE4EC"
C_LEGEND_BG = "FFEEEEEE"

# ── Schedule data ──────────────────────────────────────────────────────────────
# (day_index, hour_int): (text, fill_color)
# day 0=Jul15 … day10=Jul25
S = {
    # ── Day 0 · Jul 15 · Arrival ──────────────────────────────────────────────
    (0, 11): ("✈ נחיתה ציריך\nOS135 11:00", C_GRAY),
    (0, 12): ("🚗 לקיחת רכב\nנמל התעופה", C_BLUE),
    (0, 13): ("🚗 נסיעה למלון\n~1:45 שעות", C_ORANGE),
    (0, 14): ("🚗 נסיעה למלון\nA5/B31 → Höllental", C_ORANGE),
    (0, 15): ("🌲 Ravennaschlucht\nקניון רוונה – ישירות מהמלון", C_GREEN),
    (0, 16): ("🌲 Ravennaschlucht\nמפלים + שבילי עץ", C_GREEN),
    (0, 17): ("🌲 Ravennaschlucht\nמסלול מעגלי ~4 ק״מ", C_GREEN),
    (0, 18): ("🌲 Ravennaschlucht\nחזרה למלון", C_GREEN),

    # ── Day 1 · Jul 16 ────────────────────────────────────────────────────────
    (1,  9): ("🚗 נסיעה לטיטיזה\n~10 דק'", C_ORANGE),
    (1, 10): ("🌲 אגם טיטיזה\nשיט בסירה + טיילת", C_GREEN),
    (1, 11): ("🌲 אגם טיטיזה\nשחייה חופשית", C_GREEN),
    (1, 12): ("🌲 אגם טיטיזה\nשחייה + ארוחת צהריים", C_GREEN),
    (1, 13): ("🚗 נסיעה לפלדברג\n~15 דק'", C_ORANGE),
    (1, 14): ("🌲 הר פלדברג (1493 מ')\nFeldbergbahn רכבל + מסלול גמדים", C_GREEN),
    (1, 15): ("🌲 הר פלדברג\nמגדל תצפית + מעלית פנימית", C_GREEN),
    (1, 16): ("🌲 הר פלדברג\nירידה ברגל ~40 דק'", C_GREEN),
    (1, 17): ("🎢 Badeparadies Schwarzwald\nפארק מים – מגלשות אקסטרים", C_YELLOW),
    (1, 18): ("🎢 Badeparadies\nבריכות טרופיות + מגלשות", C_YELLOW),
    (1, 19): ("🎢 Badeparadies\nבריכות טרופיות + מגלשות", C_YELLOW),

    # ── Day 2 · Jul 17 · Europa-Park ──────────────────────────────────────────
    (2,  8): ("🚗 נסיעה לפארק אירופה\n~55 דק' (Rust)", C_ORANGE),
    (2,  9): ("🎢 Europa-Park\nכניסה + ריצה לרכבות", C_YELLOW),
    (2, 10): ("🎢 Europa-Park\nרכבות הרים + אטרקציות", C_YELLOW),
    (2, 11): ("🎢 Europa-Park\nאטרקציות + אזורי מים", C_YELLOW),
    (2, 12): ("🎢 Europa-Park\nארוחת צהריים + אטרקציות", C_YELLOW),
    (2, 13): ("🎢 Europa-Park\nרכבות הרים עיקריות", C_YELLOW),
    (2, 14): ("🎢 Europa-Park\nאזור Silver Star / Blue Fire", C_YELLOW),
    (2, 15): ("🎢 Europa-Park\nאזור Arthur + Voletarium", C_YELLOW),
    (2, 16): ("🎢 Europa-Park\nשואו ערב + אטרקציות", C_YELLOW),
    (2, 17): ("🎢 Europa-Park\nשואו ערב", C_YELLOW),
    (2, 18): ("🎢 Europa-Park\nשעות סגירה", C_YELLOW),
    (2, 19): ("🚗 נסיעה חזרה למלון\n~55 דק'", C_ORANGE),

    # ── Day 3 · Jul 18 ────────────────────────────────────────────────────────
    (3,  9): ("🚗 נסיעה לטודנאו\n~20 דק'", C_ORANGE),
    (3, 10): ("🌲 מפלי טודנאו\nTodtnauer Wasserfall", C_GREEN),
    (3, 11): ("🌲 מפלי טודנאו\nמסלול הליכה יפה בטבע", C_GREEN),
    (3, 12): ("🎢 Hasenhorn Rodelbahn\nרכבל + מגלשות הרים", C_YELLOW),
    (3, 13): ("🎢 Hasenhorn\nמגלשות + תצפית מרהיבה", C_YELLOW),
    (3, 14): ("🎢 Hasenhorn\nסיבובים נוספים", C_YELLOW),
    (3, 15): ("🎢 Blackforestline\nגשר חבלים תלוי מעל המפלים", C_YELLOW),
    (3, 16): ("🎢 Blackforestline\nגשר חבלים תלוי", C_YELLOW),

    # ── Day 4 · Jul 19 ────────────────────────────────────────────────────────
    (4,  9): ("🚗 נסיעה לשלוכזי\n~30 דק'", C_ORANGE),
    (4, 10): ("🌲 אגם שלוכזי\nקייאק / SUP – אגם הגדול ביותר", C_GREEN),
    (4, 11): ("🌲 אגם שלוכזי\nשחייה + שיטים", C_GREEN),
    (4, 12): ("🌲 אגם שלוכזי\nארוחת צהריים + שחייה", C_GREEN),
    (4, 13): ("🚗 נסיעה ל-Action Forest\n~10 דק' (Titisee)", C_ORANGE),
    (4, 14): ("🎢 Action Forest\nפארק חבלים – 6 מסלולים", C_YELLOW),
    (4, 15): ("🎢 Action Forest\nזיפ-ליין + גובה 21 מ'", C_YELLOW),
    (4, 16): ("🎢 Action Forest\nמסלולי אתגר בגובה", C_YELLOW),

    # ── Day 5 · Jul 20 ────────────────────────────────────────────────────────
    (5,  9): ("🚗 נסיעה לטריברג\n~45 דק' (צפונה)", C_ORANGE),
    (5, 10): ("🌲 מפלי טריברג\nהמפלים הגבוהים בגרמניה", C_GREEN),
    (5, 11): ("🌲 מפלי טריברג\nהאכלת סנאים + מסלול יער", C_GREEN),
    (5, 12): ("🌲 טריברג\nשעוני קוקייה + ארוחת צהריים", C_GREEN),
    (5, 13): ("🚗 נסיעה לגוטאך\n~15 דק'", C_ORANGE),
    (5, 14): ("🎢 Vogtsbauernhof\nמוזיאון חוות פתוח – גוטאך", C_YELLOW),
    (5, 15): ("🎢 Vogtsbauernhof\n+ Sommerrodelbahn מגלשות", C_YELLOW),
    (5, 16): ("🎢 Gutach\nמגלשות הרים – סיבובים", C_YELLOW),

    # ── Day 6 · Jul 21 · Transition ───────────────────────────────────────────
    (6,  9): ("🏨 צ'ק אאוט\nHofgut Sternen", C_GRAY),
    (6, 10): ("🌲 Wutachschlucht\nגרנד קניון השחורצוואלד", C_GREEN),
    (6, 11): ("🌲 Wutachschlucht\nמסלול לאורך הנהר הפראי", C_GREEN),
    (6, 12): ("🌲 Wutachschlucht\nמפלים + נקיקים (15 ק״מ)", C_GREEN),
    (6, 13): ("🌲 Wutachschlucht\nאוטובוס חזרה לרכב", C_GREEN),
    (6, 14): ("🎢 Tannenmühle\nחווה + ארוחה + בריכות דגים", C_YELLOW),
    (6, 15): ("🎢 Tannenmühle\nחיות + בריכות", C_YELLOW),
    (6, 16): ("🎢 Tannenmühle\nמרכז מבקרים", C_YELLOW),
    (6, 17): ("🚗 נסיעה למלון החדש\nWaldshut-Tiengen ~30 דק'", C_ORANGE),

    # ── Day 7 · Jul 22 · Rhine Falls + Stein am Rhein ────────────────────────
    (7,  9): ("🚗 נסיעה למפלי הריין\nSchaffhausen, שוויץ ~30 דק'", C_ORANGE),
    (7, 10): ("🎢 Rhine Falls\nשיט לסלע המרכזי", C_YELLOW),
    (7, 11): ("🎢 Rhine Falls\nטירת לאופן + נקודות תצפית", C_YELLOW),
    (7, 12): ("🎢 Rhine Falls\nAdventure Park חבלים", C_YELLOW),
    (7, 13): ("🚗 נסיעה לשטיין אם ריין\n~20 דק'", C_ORANGE),
    (7, 14): ("🌲 Stein am Rhein\nעיירה מימי ביניים – ציורי קיר", C_GREEN),
    (7, 15): ("🌲 Stein am Rhein\nטיילת על נהר הריין", C_GREEN),
    (7, 16): ("🌲 Stein am Rhein\nסיור חופשי + גלידה", C_GREEN),

    # ── Day 8 · Jul 23 · Albsteig + St. Blasien ──────────────────────────────
    (8,  9): ("🚗 נסיעה ל-Görwihl\n~20 דק'", C_ORANGE),
    (8, 10): ("🌲 Albsteig + מפלי Höllbach\nמסלול מעגלי פראי – 6 ק״מ", C_GREEN),
    (8, 11): ("🌲 Albsteig\nמפלי הולבך + נקיקים", C_GREEN),
    (8, 12): ("🌲 Albsteig\nמסלול בטבע הפראי", C_GREEN),
    (8, 13): ("🌲 Albsteig\nארוחת צהריים + סיום מסלול", C_GREEN),
    (8, 14): ("🚗 נסיעה לסנט בלאזיין\n~30 דק'", C_ORANGE),
    (8, 15): ("🌲 St. Blasien\nקתדרלה עם הכיפה הגדולה", C_GREEN),
    (8, 16): ("🌲 St. Blasien\nסיור בעיירה האלפינית", C_GREEN),

    # ── Day 9 · Jul 24 · Freiburg + Schauinsland ─────────────────────────────
    (9,  9): ("🚗 נסיעה לפרייבורג\n~1:15 שעות", C_ORANGE),
    (9, 10): ("🌲 פרייבורג\nעיר עתיקה + תעלות מים (Bächle)", C_GREEN),
    (9, 11): ("🌲 פרייבורג\nמינסטר + שוק + קתדרלה", C_GREEN),
    (9, 12): ("🌲 פרייבורג\nארוחת צהריים בעיר", C_GREEN),
    (9, 13): ("🌲 פרייבורג\nסיור חופשי + קניות", C_GREEN),
    (9, 14): ("🚗 נסיעה לשקאונסלנד\n~15 דק'", C_ORANGE),
    (9, 15): ("🎢 Schauinslandbahn\nרכבל ארוך + תצפית", C_YELLOW),
    (9, 16): ("🎢 Schauinslandbahn\nסיור במכרה עתיק", C_YELLOW),
    (9, 17): ("🎢 Schauinslandbahn\nתצפיות על העמק", C_YELLOW),

    # ── Day 10 · Jul 25 · Return ──────────────────────────────────────────────
    (10,  9): ("🌲 וולדסהוט-טיינגן\nמדרחוב היסטורי", C_GREEN),
    (10, 10): ("🌲 וולדסהוט-טיינגן\nטיילת ריין + קניות", C_GREEN),
    (10, 11): ("🌲 וולדסהוט-טיינגן\nארוחת צהריים + פרידה", C_GREEN),
    (10, 13): ("📦 ארגון מזוודות\nהכנה לנסיעה", C_GRAY),
    (10, 14): ("📦 ארגון אחרון\nאריזה ופינוי חדר", C_GRAY),
    (10, 15): ("📦 סיכום + ארגון\nלפני יציאה", C_GRAY),
    (10, 16): ("🚗 נסיעה לנמל התעופה\nציריך ~45-60 דק'", C_ORANGE),
    (10, 17): ("🚗 בדרך לציריך\nנמל התעופה", C_ORANGE),
    (10, 18): ("🚗 החזרת רכב\nZurich Airport", C_BLUE),
    (10, 19): ("✈ צ'ק אין + ביטחון\nZurich Airport", C_GRAY),
    (10, 20): ("✈ המראה 20:50\nOS146 → תל אביב 🏠", C_GRAY),
    (10, 21): ("✈ טיסה OS146\nבדרך הביתה", C_GRAY),
    (10, 22): ("✈ טיסה OS146\n~3:30 שעות טיסה", C_GRAY),
}

DAYS = [
    {"date": "15/07", "dow": "יום ג'", "phase": 1},
    {"date": "16/07", "dow": "יום ד'", "phase": 1},
    {"date": "17/07", "dow": "יום ה'", "phase": 1},
    {"date": "18/07", "dow": "יום ו'", "phase": 1},
    {"date": "19/07", "dow": "שבת",    "phase": 1},
    {"date": "20/07", "dow": "יום א'", "phase": 1},
    {"date": "21/07", "dow": "יום ב'", "phase": 2},
    {"date": "22/07", "dow": "יום ג'", "phase": 2},
    {"date": "23/07", "dow": "יום ד'", "phase": 2},
    {"date": "24/07", "dow": "יום ה'", "phase": 2},
    {"date": "25/07", "dow": "יום ו'", "phase": 3},
]

HOURS = list(range(7, 24))  # 07:00 – 23:00

THIN = border()

wb = Workbook()
ws = wb.active
ws.title = "Black Forest 2026"
ws.sheet_view.rightToLeft = True

# ── Column widths ──
ws.column_dimensions['A'].width = 7
for i in range(2, 13):
    ws.column_dimensions[get_column_letter(i)].width = 27

# ── Row heights ──
ws.row_dimensions[1].height = 38
ws.row_dimensions[2].height = 28
ws.row_dimensions[3].height = 24
for r in range(4, 4 + len(HOURS)):
    ws.row_dimensions[r].height = 52

# ── ROW 1 – Title ──
ws.merge_cells('A1:L1')
c = ws['A1']
c.value = "🌲  Black Forest & Rhine  ·  Family Trip July 2026  ·  Guy, עדי, עמית, נדב, איתי  🌲"
c.fill = fill(C_HDR_DARK)
c.font = Font(name='Arial', bold=True, color="FFFFFFFF", size=13)
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = THIN

# ── ROW 2 – Hotel banners ──
# Time label cell
c = ws['A2']
c.fill = fill("FF263238"); c.border = THIN

# Hotel 1: B2:G2 (days 0-5, Jul 15-20)
ws.merge_cells('B2:G2')
c = ws['B2']
c.value = "🏨  Hotel Hofgut Sternen · Breitnau · Black Forest  (15/07 – 21/07)"
c.fill = fill(C_HDR_H1)
c.font = Font(name='Arial', bold=True, color="FFFFFFFF", size=11)
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = THIN

# Hotel 2: H2:K2 (days 6-9, Jul 21-24)
ws.merge_cells('H2:K2')
c = ws['H2']
c.value = "🏨  Waldshut-Tiengen  (21/07 – 25/07)"
c.fill = fill(C_HDR_H2)
c.font = Font(name='Arial', bold=True, color="FFFFFFFF", size=11)
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = THIN

# Return day: L2
c = ws['L2']
c.value = "✈️  Return"
c.fill = fill(C_HDR_END)
c.font = Font(name='Arial', bold=True, color="FFFFFFFF", size=11)
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = THIN

# ── ROW 3 – Day headers ──
c = ws['A3']
c.value = "שעה"
c.fill = fill("FF37474F")
c.font = Font(name='Arial', bold=True, color="FFFFFFFF", size=10)
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = THIN

DAY_HDR_COLORS = {1: C_DAY_H1, 2: C_DAY_H2, 3: C_DAY_END}
for di, day in enumerate(DAYS):
    col = get_column_letter(di + 2)
    c = ws[f"{col}3"]
    c.value = f"{day['date']}\n{day['dow']}"
    c.fill = fill(DAY_HDR_COLORS[day['phase']])
    c.font = Font(name='Arial', bold=True, color="FFFFFFFF", size=10)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = THIN

# ── ROWS 4+ – Time slots ──
EMPTY_COLORS = {1: C_EMPTY_H1, 2: C_EMPTY_H2, 3: C_EMPTY_END}

for hi, hour in enumerate(HOURS):
    row = hi + 4
    # Time label
    c = ws[f"A{row}"]
    c.value = f"{hour:02d}:00"
    c.fill = fill("FF546E7A")
    c.font = Font(name='Arial', bold=True, color="FFFFFFFF", size=10)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = THIN

    for di, day in enumerate(DAYS):
        col = get_column_letter(di + 2)
        c = ws[f"{col}{row}"]
        key = (di, hour)
        if key in S:
            text, bg = S[key]
            c.value = text
            c.fill = fill(bg)
            c.font = Font(name='Arial', size=8)
        else:
            c.value = ""
            c.fill = fill(EMPTY_COLORS[day['phase']])
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = THIN

# ── Legend ──
leg_row = 4 + len(HOURS) + 1
ws.merge_cells(f'A{leg_row}:B{leg_row}')
c = ws[f'A{leg_row}']
c.value = "מקרא:"
c.fill = fill(C_LEGEND_BG)
c.font = Font(name='Arial', bold=True, size=10)
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = THIN

LEGENDS = [
    ("✈  טיסות / שדה תעופה / ארגון",       C_GRAY),
    ("🚗  לקיחת / החזרת רכב",               C_BLUE),
    ("🚗  נסיעות ברכב",                      C_ORANGE),
    ("🌲  טיול טבע / הליכה / אגמים",         C_GREEN),
    ("🎢  אטרקציות / פארקים / הרפתקאות",     C_YELLOW),
]
for i, (txt, bg) in enumerate(LEGENDS):
    col_s = get_column_letter(i * 2 + 1)
    col_e = get_column_letter(i * 2 + 2)
    if i * 2 + 2 <= 12:
        ws.merge_cells(f'{col_s}{leg_row+1}:{col_e}{leg_row+1}')
    c = ws[f'{col_s}{leg_row+1}']
    c.value = txt
    c.fill = fill(bg)
    c.font = Font(name='Arial', bold=True, size=9)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = THIN

# ── Save ──
out = r"D:\code\Vacations\BlackForest\Black_Forest_Trip_2026.xlsx"
wb.save(out)
print(f"Saved: {out}")
